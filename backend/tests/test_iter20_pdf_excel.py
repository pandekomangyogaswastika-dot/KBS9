"""
Iteration 20: Test PDF export (all options), Excel export (_Opts sheet), Excel import, CMS PDF config
"""
import pytest
import requests
import os
import tempfile

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

SUBMITTED_SESSION = "912ff0c5-41d4-458d-b4dc-d7cefd35561f"
DRAFT_SESSION = "81d20dfe-71ca-4d3c-aef9-d2655683fb3d"

ADMIN_EMAIL = "admin@kubus.id"
ADMIN_PASSWORD = "Admin#2026"


@pytest.fixture(scope="module")
def admin_token():
    resp = requests.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    assert resp.status_code == 200, f"Login failed: {resp.text}"
    return resp.json()["data"]["access_token"]


@pytest.fixture(scope="module")
def auth_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


class TestPDFExport:
    """PDF Export tests"""

    def test_pdf_export_submitted_session_status(self, auth_headers):
        resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-pdf", headers=auth_headers)
        assert resp.status_code == 200, f"PDF export failed: {resp.status_code} {resp.text[:200]}"
        print(f"PDF export submitted: status={resp.status_code}, size={len(resp.content)} bytes")

    def test_pdf_export_submitted_session_size(self, auth_headers):
        resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-pdf", headers=auth_headers)
        assert len(resp.content) > 3000, f"PDF too small: {len(resp.content)} bytes"
        print(f"PDF size: {len(resp.content)} bytes - PASS")

    def test_pdf_export_submitted_content_type(self, auth_headers):
        resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-pdf", headers=auth_headers)
        assert "pdf" in resp.headers.get("content-type", "").lower(), f"Wrong content-type: {resp.headers.get('content-type')}"

    def test_pdf_export_draft_session(self, auth_headers):
        resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{DRAFT_SESSION}/export-pdf", headers=auth_headers)
        assert resp.status_code == 200, f"Draft PDF export failed: {resp.status_code} {resp.text[:200]}"
        print(f"Draft PDF size: {len(resp.content)} bytes")


class TestExcelExportImport:
    """Excel Export and Import tests"""

    def test_excel_export_status(self, auth_headers):
        resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-answers-excel", headers=auth_headers)
        assert resp.status_code == 200, f"Excel export failed: {resp.status_code} {resp.text[:200]}"
        print(f"Excel export: status={resp.status_code}, size={len(resp.content)} bytes")

    def test_excel_export_content_type(self, auth_headers):
        resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-answers-excel", headers=auth_headers)
        ct = resp.headers.get("content-type", "")
        assert "spreadsheet" in ct or "excel" in ct or "octet" in ct, f"Unexpected content-type: {ct}"

    def test_excel_has_opts_sheet(self, auth_headers):
        """Verify _Opts sheet exists in exported Excel"""
        import openpyxl
        import io
        resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-answers-excel", headers=auth_headers)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        sheet_names = wb.sheetnames
        print(f"Sheet names: {sheet_names}")
        assert "_Opts" in sheet_names, f"_Opts sheet not found. Sheets: {sheet_names}"
        print("_Opts sheet found - PASS")

    def test_excel_opts_sheet_has_labels(self, auth_headers):
        """Verify _Opts sheet has content (full labels)"""
        import openpyxl
        import io
        resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-answers-excel", headers=auth_headers)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        opts_sheet = wb["_Opts"]
        # Should have at least one row with data
        rows = list(opts_sheet.iter_rows(values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        print(f"_Opts rows with data: {len(non_empty)}")
        assert len(non_empty) > 0, "_Opts sheet is empty"
        print(f"_Opts first row: {rows[0] if rows else 'empty'}")

    def test_excel_import(self, auth_headers):
        """Export then re-import the same Excel file"""
        # Export first
        resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-answers-excel", headers=auth_headers)
        assert resp.status_code == 200
        excel_bytes = resp.content

        # Import back
        files = {"file": ("answers.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        import_resp = requests.post(
            f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/import-answers-excel",
            headers={"Authorization": auth_headers["Authorization"]},
            files=files
        )
        assert import_resp.status_code == 200, f"Import failed: {import_resp.status_code} {import_resp.text[:300]}"
        print(f"Import response: {import_resp.json()}")


class TestCMSPDFConfig:
    """CMS PDF Config tests"""

    def test_get_pdf_config(self, auth_headers):
        resp = requests.get(f"{BASE_URL}/api/admin/pdf-config", headers=auth_headers)
        assert resp.status_code == 200, f"GET pdf-config failed: {resp.text}"
        data = resp.json()
        print(f"PDF config keys: {list(data.keys())}")

    def test_put_pdf_config(self, auth_headers):
        # Get current config first
        resp = requests.get(f"{BASE_URL}/api/admin/pdf-config", headers=auth_headers)
        config = resp.json()
        config["company_name"] = "TEST Kubus 2026 Iter20"
        put_resp = requests.put(f"{BASE_URL}/api/admin/pdf-config", json=config, headers=auth_headers)
        assert put_resp.status_code == 200, f"PUT pdf-config failed: {put_resp.text}"
        print(f"PUT pdf-config: {put_resp.status_code}")

    def test_pdf_preview(self, auth_headers):
        resp = requests.get(
            f"{BASE_URL}/api/admin/pdf-config/preview",
            params={"session_id": SUBMITTED_SESSION},
            headers=auth_headers
        )
        assert resp.status_code == 200, f"PDF preview failed: {resp.status_code} {resp.text[:200]}"
        assert len(resp.content) > 0, "PDF preview returned empty content"
        print(f"PDF preview size: {len(resp.content)} bytes")
