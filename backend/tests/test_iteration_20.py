"""
Iteration 20: Test PDF Export (enhanced options display) + Excel Export/Import + CMS PDF Settings
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://farm-stack-demo-1.preview.emergentagent.com').rstrip('/')

SUBMITTED_SESSION = "912ff0c5-41d4-458d-b4dc-d7cefd35561f"
DRAFT_SESSION = "81d20dfe-71ca-4d3c-aef9-d2655683fb3d"

@pytest.fixture(scope="module")
def auth_headers():
    resp = requests.post(f"{BASE_URL}/api/auth/login", json={"email": "admin@kubus.id", "password": "Admin#2026"})
    assert resp.status_code == 200, f"Login failed: {resp.text}"
    token = resp.json().get("data", {}).get("access_token") or resp.json().get("access_token")
    return {"Authorization": f"Bearer {token}"}


# PDF Export - submitted session
def test_pdf_export_submitted(auth_headers):
    resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-pdf", headers=auth_headers)
    assert resp.status_code == 200, f"PDF export failed: {resp.status_code} {resp.text[:300]}"
    assert len(resp.content) > 3000, f"PDF too small: {len(resp.content)} bytes"
    print(f"PDF size: {len(resp.content)} bytes - PASS")


# PDF Export - draft session
def test_pdf_export_draft(auth_headers):
    resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{DRAFT_SESSION}/export-pdf", headers=auth_headers)
    assert resp.status_code == 200, f"PDF draft export failed: {resp.status_code} {resp.text[:300]}"
    assert len(resp.content) > 3000, f"PDF draft too small: {len(resp.content)} bytes"
    print(f"PDF draft size: {len(resp.content)} bytes - PASS")


# Excel Export
def test_excel_export(auth_headers):
    resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-answers-excel", headers=auth_headers)
    assert resp.status_code == 200, f"Excel export failed: {resp.status_code} {resp.text[:300]}"
    assert len(resp.content) > 5000, f"Excel too small: {len(resp.content)} bytes"
    print(f"Excel size: {len(resp.content)} bytes - PASS")


# Excel content check - _Opts sheet and Panduan sheet
def test_excel_sheets(auth_headers):
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")
    
    resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-answers-excel", headers=auth_headers)
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    sheet_names = wb.sheetnames
    print(f"Sheets found: {sheet_names}")
    
    assert "Jawaban" in sheet_names, f"'Jawaban' sheet missing. Found: {sheet_names}"
    assert "_Opts" in sheet_names, f"'_Opts' sheet missing. Found: {sheet_names}"
    assert "Panduan" in sheet_names, f"'Panduan' sheet missing. Found: {sheet_names}"
    
    # Check _Opts has content
    opts_sheet = wb["_Opts"]
    opts_rows = list(opts_sheet.iter_rows(values_only=True))
    assert len(opts_rows) > 1, "_Opts sheet has no data"
    print(f"_Opts rows: {len(opts_rows)} - PASS")
    
    # Check Jawaban has data
    jaw_sheet = wb["Jawaban"]
    jaw_rows = list(jaw_sheet.iter_rows(values_only=True))
    assert len(jaw_rows) > 1, "Jawaban sheet has no data"
    print(f"Jawaban rows: {len(jaw_rows)} - PASS")


# Excel Import
def test_excel_import(auth_headers):
    # Download first
    resp = requests.get(f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/export-answers-excel", headers=auth_headers)
    assert resp.status_code == 200
    
    # Upload back
    files = {"file": ("answers.xlsx", io.BytesIO(resp.content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    upload_resp = requests.post(
        f"{BASE_URL}/api/assessment/sessions/{SUBMITTED_SESSION}/import-answers-excel",
        headers={"Authorization": auth_headers["Authorization"]},
        files=files
    )
    # Submitted session returns 409 (business rule); accept both 200 and 409 ALREADY_SUBMITTED
    if upload_resp.status_code == 409:
        err_code = upload_resp.json().get("error", {}).get("code", "")
        assert err_code == "ALREADY_SUBMITTED", f"Unexpected 409: {upload_resp.text[:300]}"
        print(f"Import blocked for submitted session (expected behavior): {err_code} - NOTE")
    else:
        assert upload_resp.status_code == 200, f"Import failed: {upload_resp.status_code} {upload_resp.text[:300]}"
        print(f"Excel import response: {upload_resp.json()} - PASS")


# CMS PDF Config GET
def test_pdf_config_get(auth_headers):
    resp = requests.get(f"{BASE_URL}/api/admin/pdf-config", headers=auth_headers)
    assert resp.status_code == 200, f"PDF config GET failed: {resp.status_code} {resp.text[:300]}"
    resp_json = resp.json()
    data = resp_json.get("data", resp_json)
    assert "company_name" in data, f"company_name missing in response: {data}"
    print(f"PDF config: {data} - PASS")


# CMS PDF Config PUT
def test_pdf_config_put(auth_headers):
    # GET current
    get_resp = requests.get(f"{BASE_URL}/api/admin/pdf-config", headers=auth_headers)
    current = get_resp.json()
    
    update_data = {**current, "company_name": "TEST Kubus Teknologi"}
    put_resp = requests.put(f"{BASE_URL}/api/admin/pdf-config", headers=auth_headers, json=update_data)
    assert put_resp.status_code == 200, f"PUT failed: {put_resp.status_code} {put_resp.text[:300]}"
    
    # Restore
    restore_resp = requests.put(f"{BASE_URL}/api/admin/pdf-config", headers=auth_headers, json=current)
    assert restore_resp.status_code == 200
    print(f"PDF config PUT and restore - PASS")


# CMS PDF Preview
def test_pdf_config_preview(auth_headers):
    resp = requests.get(
        f"{BASE_URL}/api/admin/pdf-config/preview",
        params={"session_id": SUBMITTED_SESSION, "locale": "id"},
        headers=auth_headers
    )
    assert resp.status_code == 200, f"PDF preview failed: {resp.status_code} {resp.text[:300]}"
    assert len(resp.content) > 3000, f"Preview PDF too small: {len(resp.content)} bytes"
    print(f"PDF preview size: {len(resp.content)} bytes - PASS")
