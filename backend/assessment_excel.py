"""Excel import/export helpers for Assessment Templates.

Excel format (single sheet "Pertanyaan"):
| section_name_id | section_name_en | section_color | question_text_id | question_text_en | type | required | options_id | options_en | hint_id | hint_en |
|-----------------|-----------------|---------------|------------------|------------------|------|----------|------------|------------|---------|---------|

- section_name_id/en: Group questions by section (rows with same name belong to same section)
- options_id/en: Pipe-separated "Opsi 1|Opsi 2" (only for select/multiselect)
- type: text | textarea | select | multiselect | yesno | scale | number
- required: yes/no (default: yes)
"""
import io
from uuid import uuid4

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

QUESTION_TYPES = {"text", "textarea", "select", "multiselect", "yesno", "scale", "number"}
HEADER_COLS = [
    "section_name_id", "section_name_en", "section_color",
    "question_text_id", "question_text_en",
    "type", "required", "weight",
    "options_id", "options_en",
    "hint_id", "hint_en",
]

SAMPLE_ROWS = [
    ["Profil Perusahaan", "Company Profile", "#5B49C9",
     "Apa jenis bisnis utama perusahaan Anda?", "What is your company's main business type?",
     "select", "yes", "1",
     "Manufaktur|Distribusi|Retail|Jasa", "Manufacturing|Distribution|Retail|Services",
     "Pilih yang paling mendekati.", "Choose the closest match."],
    ["Profil Perusahaan", "Company Profile", "#5B49C9",
     "Berapa total karyawan saat ini?", "How many total employees do you have?",
     "select", "yes", "1",
     "1-25|26-100|101-250|251+", "1-25|26-100|101-250|251+",
     "", ""],
    ["Profil Perusahaan", "Company Profile", "#5B49C9",
     "Apakah ada rencana ekspansi dalam 2 tahun ke depan?", "Do you have expansion plans in the next 2 years?",
     "yesno", "yes", "1", "", "", "", ""],
    ["Infrastruktur IT", "IT Infrastructure", "#1D7874",
     "Sistem apa yang saat ini Anda gunakan?", "What systems are you currently using?",
     "multiselect", "yes", "1",
     "ERP|CRM|Akuntansi Manual|Spreadsheet|Tidak ada", "ERP|CRM|Manual Accounting|Spreadsheet|None",
     "Pilih semua yang berlaku.", "Select all that apply."],
    ["Infrastruktur IT", "IT Infrastructure", "#1D7874",
     "Deskripsikan tantangan terbesar dalam operasional saat ini.", "Describe your biggest operational challenge.",
     "textarea", "no", "1", "", "", "Ceritakan dengan detail.", "Please describe in detail."],
]


def generate_excel_template() -> bytes:
    """Generate a downloadable Excel template for assessment import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pertanyaan"

    # Header row
    header_fill = PatternFill(start_color="1E2030", end_color="1E2030", fill_type="solid")
    header_font = Font(bold=True, color="C8C6F7", size=10)
    for ci, col in enumerate(HEADER_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Column widths
    widths = [20, 20, 14, 40, 40, 12, 10, 8, 35, 35, 25, 25]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # Sample rows
    row_fill_a = PatternFill(start_color="12141E", end_color="12141E", fill_type="solid")
    row_fill_b = PatternFill(start_color="0D0F17", end_color="0D0F17", fill_type="solid")
    row_font = Font(color="E0DFF0", size=9)
    for ri, row in enumerate(SAMPLE_ROWS, 2):
        fill = row_fill_a if ri % 2 == 0 else row_fill_b
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.font = row_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 32

    # Second sheet: instructions
    ws2 = wb.create_sheet("Panduan")
    instructions = [
        ["Field", "Keterangan"],
        ["section_name_id", "Nama domain/seksi (Bahasa Indonesia). Baris dengan nama sama = 1 domain."],
        ["section_name_en", "Nama domain/seksi (Bahasa Inggris). Boleh kosong."],
        ["section_color", "Warna hex domain, contoh: #5B49C9. Boleh kosong."],
        ["question_text_id", "Teks pertanyaan (WAJIB, Bahasa Indonesia)."],
        ["question_text_en", "Teks pertanyaan (Bahasa Inggris). Boleh kosong."],
        ["type", "Tipe pertanyaan: text | textarea | select | multiselect | yesno | scale | number"],
        ["required", "yes atau no (default: yes)"],
        ["weight", "Bobot angka, biasanya 1 (default: 1)"],
        ["options_id", "Untuk select/multiselect: pisahkan opsi dengan | contoh: Opsi1|Opsi2|Opsi3"],
        ["options_en", "Terjemahan opsi (urutan sama): Option1|Option2|Option3"],
        ["hint_id", "Teks bantuan (opsional, Bahasa Indonesia)"],
        ["hint_en", "Teks bantuan (opsional, Bahasa Inggris)"],
        [],
        ["Catatan", "Kosongkan baris untuk memisahkan domain. Hapus baris contoh sebelum import."],
    ]
    for ri, row in enumerate(instructions, 1):
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if ri == 1:
                cell.font = Font(bold=True, color="C8C6F7")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def parse_excel_template(file_bytes: bytes) -> list[dict]:
    """Parse Excel file and return list of section dicts with questions."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    # Find the questions sheet (first sheet or 'Pertanyaan')
    ws = None
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ("pertanyaan", "questions", "sheet"):
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Detect header row
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col_map = {}
    for expected in HEADER_COLS:
        try:
            col_map[expected] = header.index(expected)
        except ValueError:
            pass  # optional columns

    def get(row, key, default=""):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return default
        v = row[idx]
        return str(v).strip() if v is not None else default

    sections: dict[str, dict] = {}
    section_order: list[str] = []

    for row in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # skip empty rows
        q_text_id = get(row, "question_text_id")
        if not q_text_id:
            continue

        sec_id_name = get(row, "section_name_id") or "Umum"
        if sec_id_name not in sections:
            sections[sec_id_name] = {
                "id": str(uuid4()),
                "title": {
                    "id": sec_id_name,
                    "en": get(row, "section_name_en") or sec_id_name,
                },
                "color": get(row, "section_color") or "#5B49C9",
                "questions": [],
            }
            section_order.append(sec_id_name)

        q_type = get(row, "type", "text").lower()
        if q_type not in QUESTION_TYPES:
            q_type = "text"

        required_raw = get(row, "required", "yes").lower()
        required = required_raw not in ("no", "false", "0", "tidak")

        weight_raw = get(row, "weight", "1")
        try:
            weight = float(weight_raw)
        except ValueError:
            weight = 1.0

        # Parse options
        options = None
        if q_type in ("select", "multiselect"):
            opts_id = get(row, "options_id")
            opts_en = get(row, "options_en")
            if opts_id:
                id_parts = [o.strip() for o in opts_id.split("|") if o.strip()]
                en_parts = [o.strip() for o in opts_en.split("|") if o.strip()] if opts_en else id_parts
                options = []
                for i, opt_label in enumerate(id_parts):
                    en_label = en_parts[i] if i < len(en_parts) else opt_label
                    # Value = slugified version of option label
                    val = opt_label.lower().replace(" ", "_").replace("-", "_")[:30]
                    options.append({"value": val, "label": {"id": opt_label, "en": en_label}})

        hint_id = get(row, "hint_id")
        hint_en = get(row, "hint_en")
        hint = {"id": hint_id, "en": hint_en} if hint_id or hint_en else None

        question = {
            "id": str(uuid4()),
            "text": {"id": q_text_id, "en": get(row, "question_text_en") or q_text_id},
            "type": q_type,
            "required": required,
            "weight": weight,
            "options": options,
            "hint": hint,
            "show_if": None,
            "scale_labels": None,
        }
        sections[sec_id_name]["questions"].append(question)

    return [sections[k] for k in section_order]
