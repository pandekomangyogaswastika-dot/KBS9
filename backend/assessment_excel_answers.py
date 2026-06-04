"""Excel Export/Import untuk jawaban Assessment Session — Professional Edition v2.

Export: pertanyaan + opsi lengkap + jawaban saat ini → Excel formal dengan dropdown validasi.
Import: baca Excel → bulk-upsert jawaban ke sesi assessment.

Kolom: question_id (hidden) | No | Domain | Pertanyaan | Tipe | Panduan Pengisian | Jawaban Anda

Fitur v2:
  - Dropdown validasi via sheet tersembunyi _Opts (menampilkan LABEL PENUH tanpa batas 255 karakter)
  - Opsi ditampilkan LENGKAP per baris di Panduan — tidak disingkat
  - Jawaban pre-filled menampilkan label (bukan kode internal)
  - Tema profesional korporat: background putih/abu muda, header navy, teks gelap
  - Sheet Panduan: semua teks eksplisit gelap (tidak ada putih-di-putih)
"""
import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Colour palette (professional corporate) ───────────────────────────────────
HDR_BG     = "1E3A5F"   # deep navy — headers
HDR_FG     = "FFFFFF"   # white text on header
DOM_BG     = "DDE3EE"   # blue-gray for domain rows
DOM_FG     = "1A2D4F"   # dark navy for domain text
ROW_A_BG   = "FFFFFF"   # white row
ROW_B_BG   = "F5F7FB"   # very light blue-gray alt row
ROW_FG     = "1A1C25"   # almost black body text
ANS_EMPTY  = "F9F9F9"   # light gray for empty answer cell
ANS_FILLED = "EDF3F9"   # very light blue when pre-filled
HINT_FG    = "4A5568"   # medium-dark gray for hints
BORDER_CLR = "C4CBD9"   # subtle blue-gray border
ACCENT_CLR = "2C6EAB"   # blue accent for answer column header
META_BG    = "EEF2F9"   # meta rows background
WARN_BG    = "FFF3CD"   # light warning yellow for multi-choice instruction
WARN_FG    = "7A5200"   # dark amber text


# ─── Helpers ────────────────────────────────────────────────────────────────────

def _loc(val, locale="id"):
    if isinstance(val, dict):
        return val.get(locale) or val.get("id") or val.get("en") or ""
    return str(val) if val is not None else ""


def _get_domains(template: dict) -> list:
    return template.get("domains") or template.get("sections") or []


def _get_q_text(q: dict, locale: str = "id") -> str:
    field = q.get("prompt") or q.get("text") or {}
    return _loc(field, locale)


def _normalize_type(t: str) -> str:
    return {
        "select": "single_choice", "multiselect": "multi_choice",
        "yesno": "yes_no", "scale": "scale_1_5",
        "text": "text_short", "textarea": "text_long",
        "single_choice": "single_choice", "multi_choice": "multi_choice",
    }.get(t, t)


def _type_label(qtype: str) -> str:
    return {
        "single_choice": "Pilihan Tunggal",
        "multi_choice":  "Pilihan Berganda",
        "yes_no":        "Ya / Tidak",
        "scale_1_5":     "Skala 1-5",
        "text_short":    "Teks Singkat",
        "text_long":     "Teks Panjang",
        "number":        "Angka",
    }.get(_normalize_type(qtype), qtype)


def _full_options_text(q: dict, locale: str = "id") -> str:
    """Return numbered option guide — satu baris per opsi, tidak disingkat."""
    opts = q.get("options") or []
    qtype = _normalize_type(q.get("type", ""))

    if not opts:
        if qtype == "yes_no":
            return "Pilih salah satu:\n  1. Ya\n  2. Tidak"
        if qtype == "scale_1_5":
            return (
                "Isi angka 1–5:\n"
                "  1 = Awal (Initial)\n"
                "  2 = Berkembang (Developing)\n"
                "  3 = Terdefinisi (Defined)\n"
                "  4 = Terkelola (Managed)\n"
                "  5 = Optimal (Optimizing)"
            )
        if qtype in ("text_short", "text_long"):
            return "Isi dengan teks bebas"
        if qtype == "number":
            return "Isi dengan angka"
        return ""

    lines = []
    for i, o in enumerate(opts, 1):
        label = _loc(o.get("label", {}), locale) or o.get("value", "")
        lines.append(f"  {i}. {label}")

    if qtype == "multi_choice":
        header = "Pilih beberapa, pisahkan dengan tanda | :\n"
        footer = "\n  Contoh: Opsi Pertama|Opsi Kedua"
        return header + "\n".join(lines) + footer
    else:
        return "Pilih salah satu:\n" + "\n".join(lines)


def _opt_label(opt: dict, locale: str = "id") -> str:
    """Full label string for an option (used in dropdown and display)."""
    label = _loc(opt.get("label", {}), locale) or opt.get("value", "")
    return label


def _answer_to_display(q: dict, ans: dict | None, locale: str = "id") -> str:
    """Convert stored answer value → display label string for Excel cell."""
    if not ans or ans.get("skipped"):
        return ""
    val = ans.get("value")
    if val is None or val == "":
        return ""
    qtype = _normalize_type(q.get("type", ""))
    opts_by_val = {o["value"]: o for o in (q.get("options") or [])}

    if qtype == "single_choice":
        # Return LABEL (not internal value) — matches dropdown display
        opt = opts_by_val.get(str(val))
        return _opt_label(opt, locale) if opt else str(val)

    if qtype == "multi_choice":
        vals = val if isinstance(val, list) else [val]
        labels = []
        for v in vals:
            opt = opts_by_val.get(str(v))
            labels.append(_opt_label(opt, locale) if opt else str(v))
        return "|".join(labels)

    if qtype == "yes_no":
        if val in (True, "yes", "true", "Ya", "ya"):
            return "Ya"
        if val in (False, "no", "false", "Tidak", "tidak"):
            return "Tidak"
        return str(val)

    if qtype == "scale_1_5":
        return str(val)

    return str(val)


# ─── Styling helpers ─────────────────────────────────────────────────────────────

def _thin_border(color=BORDER_CLR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value,
          bg=None, fg=ROW_FG, bold=False, italic=False,
          size=9, wrap=True, halign="left", valign="top",
          border=None, name="Calibri"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=name, size=size, bold=bold, italic=italic, color=fg)
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = border if border is not None else _thin_border()
    return cell


# ─── Options sheet builder ───────────────────────────────────────────────────────

def _build_opts_sheet(wb: openpyxl.Workbook, domains: list, locale: str) -> dict:
    """Create hidden '_Opts' sheet with full option labels per question.

    Returns {qid: DataValidation} — pre-created DVs referencing the _Opts ranges.
    The caller must call ws.add_data_validation(dv) for each DV.
    """
    ws_opts = wb.create_sheet("_Opts")
    ws_opts.sheet_state = "hidden"

    dvs: dict[str, DataValidation] = {}
    col = 1

    for dom in domains:
        for q in (dom.get("questions") or []):
            qid = q.get("id", "")
            if not qid:
                continue
            qtype = _normalize_type(q.get("type", ""))
            opts = q.get("options") or []

            # Build label list for dropdown
            if qtype == "yes_no":
                labels = ["Ya", "Tidak"]
            elif qtype == "scale_1_5":
                labels = ["1", "2", "3", "4", "5"]
            elif qtype == "single_choice" and opts:
                labels = [_opt_label(o, locale) for o in opts]
            else:
                continue  # no dropdown for multi_choice / free text

            if not labels:
                continue

            col_letter = get_column_letter(col)
            for ri, lbl in enumerate(labels, 1):
                ws_opts.cell(row=ri, column=col, value=lbl)
                ws_opts.column_dimensions[col_letter].width = max(
                    ws_opts.column_dimensions[col_letter].width or 0,
                    len(str(lbl)) * 1.1,
                )

            range_ref = f"_Opts!${col_letter}$1:${col_letter}${len(labels)}"
            dv = DataValidation(
                type="list",
                formula1=range_ref,
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=False,
            )
            dvs[qid] = dv
            col += 1

    return dvs


# ─── Export ──────────────────────────────────────────────────────────────────────

# Column index constants
C_QID    = 1   # A – hidden
C_NO     = 2   # B
C_DOMAIN = 3   # C
C_Q      = 4   # D
C_TYPE   = 5   # E
C_GUIDE  = 6   # F  (Panduan Pengisian)
C_ANS    = 7   # G  (Jawaban Anda — dropdown)

COL_WIDTHS = [28, 5, 24, 50, 18, 52, 30]


def generate_answers_excel(session: dict, template: dict, answers_map: dict,
                            locale: str = "id") -> bytes:
    """Generate professional Excel assessment answer form."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jawaban"
    ws.sheet_view.showGridLines = False

    domains = _get_domains(template)

    # ── Pre-build _Opts sheet & DVs ─────────────────────────────────────────
    q_dv_map = _build_opts_sheet(wb, domains, locale)
    for dv in q_dv_map.values():
        ws.add_data_validation(dv)

    # ── Document title block ─────────────────────────────────────────────────
    tpl_name = _loc(template.get("name", {}), locale) or "Assessment"
    client   = session.get("client_name") or "—"
    project  = session.get("project_name") or "—"
    date_str = __import__("datetime").datetime.now(
        __import__("datetime").timezone.utc
    ).strftime("%d %B %Y")

    # Row 1: Document title
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1,
                         value=f"FORMULIR JAWABAN ASSESSMENT — {tpl_name.upper()}")
    title_cell.font = Font(name="Calibri", size=13, bold=True, color=HDR_FG)
    title_cell.fill = PatternFill(start_color=HDR_BG, end_color=HDR_BG, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = _thin_border()
    ws.row_dimensions[1].height = 28

    # Rows 2-3: Meta info
    meta_rows = [
        ("Klien / Client:",   client,  "Tanggal / Date:",  date_str),
        ("Proyek / Project:", project, "Status:",          (session.get("status") or "draft").upper()),
    ]
    for mi, (k1, v1, k2, v2) in enumerate(meta_rows, 2):
        ws.merge_cells(f"A{mi}:C{mi}")
        ws.merge_cells(f"D{mi}:E{mi}")
        ws.merge_cells(f"F{mi}:G{mi}")
        for col_idx, val, is_key in [(1, k1, True), (4, v1, False), (6, k2, True)]:
            c = ws.cell(row=mi, column=col_idx, value=val)
            c.font = Font(name="Calibri", size=9, bold=is_key,
                          color=HDR_BG if is_key else ROW_FG)
            c.fill = PatternFill(start_color=META_BG, end_color=META_BG, fill_type="solid")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = _thin_border()
        ws.row_dimensions[mi].height = 17

    # Row 4: spacer
    ws.row_dimensions[4].height = 5
    for ci in range(1, 8):
        c = ws.cell(row=4, column=ci)
        c.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # ── Column header (row 5) ────────────────────────────────────────────────
    HDR_ROW = 5
    hdr_labels = [
        "Question ID", "No.", "Domain / Seksi", "Pertanyaan",
        "Tipe Jawaban", "Panduan Pengisian & Opsi Tersedia",
        "Jawaban Anda  ✏",
    ]
    hdr_bgs = [HDR_BG] * 6 + [ACCENT_CLR]
    for ci, (lbl, bg) in enumerate(zip(hdr_labels, hdr_bgs), 1):
        c = ws.cell(row=HDR_ROW, column=ci, value=lbl)
        c.font = Font(name="Calibri", size=9, bold=True, color=HDR_FG)
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()
    ws.row_dimensions[HDR_ROW].height = 28

    # ── Data rows ────────────────────────────────────────────────────────────
    data_row = HDR_ROW + 1
    q_no = 0

    for dom in domains:
        dom_title = _loc(dom.get("title", {}), locale)
        questions = dom.get("questions") or []

        # Domain separator row
        ws.merge_cells(f"A{data_row}:G{data_row}")
        dc = ws.cell(row=data_row, column=1, value=f"  {dom_title}")
        dc.font = Font(name="Calibri", size=9, bold=True, color=DOM_FG)
        dc.fill = PatternFill(start_color=DOM_BG, end_color=DOM_BG, fill_type="solid")
        dc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        dc.border = Border(
            bottom=Side(style="medium", color=ACCENT_CLR),
            top=Side(style="thin", color=BORDER_CLR),
            left=Side(style="thin", color=BORDER_CLR),
            right=Side(style="thin", color=BORDER_CLR),
        )
        ws.row_dimensions[data_row].height = 18
        data_row += 1

        for q in questions:
            q_no += 1
            qid     = q.get("id", "")
            ans     = answers_map.get(qid)
            qtype   = q.get("type", "text_short")
            qtype_n = _normalize_type(qtype)
            ans_str = _answer_to_display(q, ans, locale)
            is_prefilled = bool(ans_str)

            row_bg = ROW_A_BG if q_no % 2 != 0 else ROW_B_BG
            ans_bg = ANS_FILLED if is_prefilled else ANS_EMPTY

            # Col A: question_id (hidden)
            _cell(ws, data_row, C_QID, qid,
                  bg=row_bg, fg="BBBBBB", size=7)

            # Col B: number
            _cell(ws, data_row, C_NO, q_no,
                  bg=row_bg, fg=HINT_FG, bold=True, size=9, halign="center")

            # Col C: domain (reference)
            _cell(ws, data_row, C_DOMAIN, dom_title,
                  bg=row_bg, fg=HINT_FG, size=8)

            # Col D: question text
            q_text = _get_q_text(q, locale) or "—"
            _cell(ws, data_row, C_Q, q_text,
                  bg=row_bg, fg=ROW_FG, size=9)

            # Col E: type label
            _cell(ws, data_row, C_TYPE, _type_label(qtype),
                  bg=row_bg, fg=HINT_FG, size=8, halign="center")

            # Col F: full options guide (multi-line, not truncated)
            guide = _full_options_text(q, locale)
            _cell(ws, data_row, C_GUIDE, guide,
                  bg=row_bg, fg=HINT_FG, italic=True, size=8)

            # Col G: answer cell (editable, with dropdown)
            ans_cell = ws.cell(row=data_row, column=C_ANS,
                               value=ans_str if ans_str else None)
            ans_cell.font = Font(name="Calibri", size=9,
                                 bold=bool(ans_str), color=ROW_FG)
            ans_cell.fill = PatternFill(start_color=ans_bg, end_color=ans_bg,
                                        fill_type="solid")
            ans_cell.alignment = Alignment(horizontal="left", vertical="top",
                                           wrap_text=True)
            ans_cell.border = Border(
                left=Side(style="medium", color=ACCENT_CLR),
                right=Side(style="thin", color=BORDER_CLR),
                top=Side(style="thin", color=BORDER_CLR),
                bottom=Side(style="thin", color=BORDER_CLR),
            )

            # Assign dropdown (from _Opts sheet)
            dv = q_dv_map.get(qid)
            if dv:
                dv.add(ans_cell)

            # Multi-choice hint in answer cell (no dropdown available)
            if qtype_n == "multi_choice" and not ans_str:
                hint_cell = ans_cell
                hint_cell.value = None
                # Add comment-like placeholder via italic gray text when empty

            # Row height: based on question length AND option count
            n_opts = len(q.get("options") or [])
            q_lines = max(1, len(q_text) // 52 + 1)
            if qtype_n in ("single_choice", "multi_choice"):
                guide_lines = n_opts + 2
            elif qtype_n == "scale_1_5":
                guide_lines = 7
            elif qtype_n == "yes_no":
                guide_lines = 3
            else:
                guide_lines = 2
            row_h = max(20, max(q_lines, guide_lines) * 12)
            ws.row_dimensions[data_row].height = row_h
            data_row += 1

    # ── Column widths & hide question_id ─────────────────────────────────────
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.column_dimensions["A"].hidden = True

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = f"D{HDR_ROW + 1}"

    # ── Print area ───────────────────────────────────────────────────────────
    ws.print_area = f"B1:G{data_row - 1}"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    # ── Panduan sheet (instructions) ─────────────────────────────────────────
    ws2 = wb.create_sheet("Panduan")
    ws2.sheet_view.showGridLines = False

    # Title row 1
    ws2.merge_cells("A1:C1")
    t = ws2.cell(row=1, column=1,
                 value="PANDUAN PENGISIAN FORMULIR JAWABAN ASSESSMENT")
    t.font = Font(name="Calibri", size=12, bold=True, color=HDR_FG)
    t.fill = PatternFill(start_color=HDR_BG, end_color=HDR_BG, fill_type="solid")
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.border = _thin_border()
    ws2.row_dimensions[1].height = 26

    # Instruction table data
    pan_rows = [
        # (Kolom, Keterangan, Catatan)  — header
        ("Kolom", "Keterangan", "Catatan"),
        ("A — Question ID",   "ID pertanyaan — JANGAN diubah. Digunakan sistem saat import.", "Wajib utuh"),
        ("B — No.",            "Nomor urut — hanya referensi, tidak diimport.", ""),
        ("C — Domain/Seksi",  "Nama domain/seksi — hanya referensi.", ""),
        ("D — Pertanyaan",    "Teks pertanyaan — hanya referensi.", ""),
        ("E — Tipe Jawaban",  "Jenis input yang diharapkan.", ""),
        ("F — Panduan",       "Daftar opsi tersedia / format pengisian — hanya referensi.", ""),
        ("G — Jawaban Anda",  "ISI KOLOM INI. Sel berwarna biru = harus diisi.", "← EDIT DI SINI"),
        ("", "", ""),
        ("Panduan per Tipe Jawaban", "Format Pengisian yang Benar", "Keterangan"),
        ("Pilihan Tunggal",    "Pilih SATU opsi dari kolom Panduan. Dropdown tersedia.", "Dropdown aktif"),
        ("Pilihan Berganda",   "Ketik beberapa opsi, pisahkan dengan |   Contoh: Opsi A|Opsi B", "Tanpa dropdown"),
        ("Ya / Tidak",         "Pilih: Ya  ATAU  Tidak", "Dropdown aktif"),
        ("Skala 1-5",          "Pilih angka 1, 2, 3, 4, atau 5", "Dropdown aktif"),
        ("Teks Singkat",       "Tulis teks bebas (1 baris).", ""),
        ("Teks Panjang",       "Tulis penjelasan, enter diperbolehkan.", ""),
        ("Angka",              "Isi dengan angka. Desimal gunakan titik: 3.5", ""),
        ("", "", ""),
        ("Catatan Penting", "Detail", ""),
        ("Baris kosong di kolom G", "Pertanyaan dianggap belum dijawab — diabaikan saat import.", ""),
        ("Jangan tambah/hapus baris", "Sistem membaca berdasarkan Question ID di kolom A.", ""),
        ("Jangan ubah nama sheet", "Sheet 'Jawaban' harus tetap bernama 'Jawaban'.", ""),
        ("Kolom A tersembunyi", "Jangan menghapus atau menggeser kolom A.", ""),
    ]

    # Section row indices (1-based within pan_rows list, starting from row 2 in sheet)
    section_rows = {2, 10, 19}   # header of each table block

    for ri, row_data in enumerate(pan_rows, 2):
        is_hdr = ri in section_rows
        is_blank = all(v == "" for v in row_data)
        bg = HDR_BG if is_hdr else ("FFFFFF" if is_blank else "FFFFFF")
        fg_color = HDR_FG if is_hdr else ROW_FG  # always dark on non-header rows

        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(row=ri, column=ci, value=val if val else None)
            c.font = Font(name="Calibri", size=9, bold=is_hdr, color=fg_color)
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=True, indent=1)
            c.border = Border(bottom=Side(style="thin", color=BORDER_CLR))

        # Highlight "Jawaban Anda" row
        if row_data[0] == "G — Jawaban Anda":
            for ci in range(1, 4):
                c = ws2.cell(row=ri, column=ci)
                c.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                                     fill_type="solid")
                c.font = Font(name="Calibri", size=9, bold=True, color=HDR_BG)

        ws2.row_dimensions[ri].height = 18 if not is_blank else 8

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 62
    ws2.column_dimensions["C"].width = 16

    # ── Meta sheet ───────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Meta")
    from datetime import datetime, timezone
    ws3["A1"], ws3["B1"] = "session_id",    session.get("id", "")
    ws3["A2"], ws3["B2"] = "template_id",   template.get("id", "")
    ws3["A3"], ws3["B3"] = "generated_at",  datetime.now(timezone.utc).isoformat()
    ws3["A4"], ws3["B4"] = "client_name",   client
    ws3["A5"], ws3["B5"] = "project_name",  project
    ws3["A6"], ws3["B6"] = "format_version", "2"

    # ── Final output ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Import ───────────────────────────────────────────────────────────────────

def _parse_answer(q: dict, raw_val: Any) -> dict | None:
    if raw_val is None or str(raw_val).strip() == "":
        return None
    raw = str(raw_val).strip()
    qtype = _normalize_type(q.get("type", "text_short"))
    opts = {o["value"]: o for o in (q.get("options") or [])}
    # Label mappings (case-insensitive)
    opts_by_label_id = {
        (_loc(o.get("label", {}), "id") or o.get("value", "")).lower(): o["value"]
        for o in (q.get("options") or [])
    }
    opts_by_label_en = {
        (_loc(o.get("label", {}), "en") or o.get("value", "")).lower(): o["value"]
        for o in (q.get("options") or [])
    }

    if qtype == "single_choice":
        # Try as value first (backwards compat), then as label
        if raw in opts:
            return {"value": raw, "skipped": False}
        lower = raw.lower()
        if lower in opts_by_label_id:
            return {"value": opts_by_label_id[lower], "skipped": False}
        if lower in opts_by_label_en:
            return {"value": opts_by_label_en[lower], "skipped": False}
        return {"value": raw, "skipped": False}

    if qtype == "multi_choice":
        parts = [p.strip() for p in raw.split("|") if p.strip()]
        resolved = []
        for p in parts:
            if p in opts:
                resolved.append(p)
            elif p.lower() in opts_by_label_id:
                resolved.append(opts_by_label_id[p.lower()])
            elif p.lower() in opts_by_label_en:
                resolved.append(opts_by_label_en[p.lower()])
            else:
                resolved.append(p)
        return {"value": resolved, "skipped": False} if resolved else None

    if qtype == "yes_no":
        lower = raw.lower()
        if lower in ("ya", "yes", "y", "true", "1"):
            return {"value": True, "skipped": False}
        if lower in ("tidak", "no", "n", "false", "0"):
            return {"value": False, "skipped": False}
        return None

    if qtype == "scale_1_5":
        try:
            n = int(float(raw))
            if 1 <= n <= 5:
                return {"value": n, "skipped": False}
        except (ValueError, TypeError):
            pass
        return None

    if qtype == "number":
        try:
            return {"value": float(raw) if "." in raw else int(raw), "skipped": False}
        except (ValueError, TypeError):
            return None

    return {"value": raw, "skipped": False}


def parse_answers_excel(file_bytes: bytes, template: dict) -> list[dict]:
    """Parse uploaded Excel → list of {question_id, value, skipped}."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Find Jawaban sheet
    ws = None
    for name in wb.sheetnames:
        if name.lower() in ("jawaban", "answers"):
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # Build question index
    q_index: dict[str, dict] = {}
    for dom in _get_domains(template):
        for q in dom.get("questions") or []:
            if q.get("id"):
                q_index[q["id"]] = q

    # Auto-detect header row and columns
    qid_col = ans_col = None
    hdr_row_idx = 0
    for ri, row in enumerate(rows):
        cells = [str(c).strip().lower() if c else "" for c in row]
        if any("question_id" in c or c == "id" for c in cells):
            hdr_row_idx = ri
            for ci, h in enumerate(cells):
                if "question_id" in h or h == "id":
                    qid_col = ci
                if "jawaban" in h or "answer" in h:
                    ans_col = ci
            break

    if qid_col is None:
        qid_col = 0
    if ans_col is None:
        ans_col = C_ANS - 1  # fallback: col G (0-indexed = 6)

    results: list[dict] = []
    for row in rows[hdr_row_idx + 1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        if qid_col >= len(row):
            continue
        qid = str(row[qid_col] or "").strip()
        if not qid or qid not in q_index:
            continue
        raw_val = row[ans_col] if ans_col < len(row) else None
        parsed = _parse_answer(q_index[qid], raw_val)
        if parsed is None:
            continue
        results.append({
            "question_id": qid,
            "value": parsed["value"],
            "skipped": parsed.get("skipped", False),
            "other_text": None,
            "note": None,
        })

    return results
